# excel_gen.py
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from pathlib import Path
import json

def build_workbook(data, xlsx_path: Path):
    wb = Workbook()
    # Sheet1: AuditData
    ws = wb.active
    ws.title = "AuditData"
    cols = data["columns"]
    ws.append(cols)
    for r in data["rows"]:
        ws.append([r.get(c,"") for c in cols])
    # Add conditional formatting on Status column if exists
    if "Status" in cols:
        idx = cols.index("Status") + 1
        col = get_column_letter(idx)
        red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        ws.conditional_formatting.add(f"{col}2:{col}{ws.max_row}", CellIsRule(operator='equal', formula=['"Non-Compliant"'], stopIfTrue=True, fill=red))
        ws.conditional_formatting.add(f"{col}2:{col}{ws.max_row}", CellIsRule(operator='equal', formula=['"Observation"'], stopIfTrue=True, fill=yellow))
        ws.conditional_formatting.add(f"{col}2:{col}{ws.max_row}", CellIsRule(operator='equal', formula=['"Compliant"'], stopIfTrue=True, fill=green))
    # Sheet2: Summary dashboard (basic formulas)
    ws2 = wb.create_sheet("Dashboard")
    ws2["A1"] = "Compliance %"
    # formula example: =COUNTIF(AuditData!C:C,"Compliant")/COUNTA(AuditData!C:C)
    if "Status" in cols:
        status_col = cols.index("Status") + 1
        col_letter = get_column_letter(status_col)
        ws2["A2"] = f"=COUNTIF(AuditData!{col_letter}:{col_letter},\"Compliant\")/COUNTA(AuditData!{col_letter}:{col_letter})"
    # Sheet3: AuditTrail
    ws3 = wb.create_sheet("AuditTrail")
    ws3.append(["Timestamp","User","Action","Details","PrevHash","Hash","Signature"])
    # write initial genesis row (optional)
    wb.save(xlsx_path)

def append_audit_footer(xlsx_path: Path, final_hash_hex: str, signature_b64: str, ts: str):
    wb = load_workbook(xlsx_path)
    if "AuditTrail" not in wb.sheetnames:
        ws = wb.create_sheet("AuditTrail")
        ws.append(["Timestamp","User","Action","Details","PrevHash","Hash","Signature"])
    else:
        ws = wb["AuditTrail"]
    # write audit footer row (FinalHash + signature)
    ws.append([ts,"System","FinalHash","Chain Root","", final_hash_hex, signature_b64])
    # optionally lock this area later (handled by macros)
    wb.save(xlsx_path)
